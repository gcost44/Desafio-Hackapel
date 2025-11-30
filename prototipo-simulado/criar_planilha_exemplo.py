"""
🏥 CRIADOR DE PLANILHA DE AGENDAMENTOS SUS
Sistema Hackapel 2025

Estrutura da planilha:
- clinica: Nome da clínica/UBS
- exame: Tipo de especialidade
- data: Data da consulta (DD/MM/YYYY)
- horario: Horário da consulta
- disponivel: SIM (livre) ou NAO (ocupado)
- paciente: Nome do paciente (vazio se disponível)
- telefone: Telefone do paciente (formato: DDD+número, ex: 53991189715)
- status_confirmacao: PENDENTE, CONFIRMADO ou CANCELADO
"""

import pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def criar_planilha_agendamentos(nome_arquivo="agenda_clinicas.xlsx"):
    """Cria planilha de agendamentos com estrutura otimizada para WhatsApp"""
    
    print("=" * 60)
    print("🏥 CRIANDO PLANILHA DE AGENDAMENTOS SUS")
    print("=" * 60)
    
    # Configurações
    clinicas = [
        "Hospital Central",
        "UBS Norte", 
        "UBS Sul",
        "Clínica Popular",
        "Centro de Saúde"
    ]
    
    exames = [
        "Cardiologista",
        "Oncologista", 
        "Ortopedista",
        "Oftalmologista",
        "Neurologista",
        "Dermatologista",
        "Nutricionista"
    ]
    
    horarios = ["08:00", "09:00", "10:00", "11:00", "14:00", "15:00", "16:00", "17:00"]
    
    # Criar dados
    data = []
    
    for dia in range(30):  # 30 dias de agenda
        data_exame = datetime.now() + timedelta(days=dia)
        data_formatada = data_exame.strftime("%d/%m/%Y")
        
        for clinica in clinicas:
            for exame in exames:
                for horario in horarios:
                    data.append({
                        "clinica": clinica,
                        "exame": exame,
                        "data": data_formatada,
                        "horario": horario,
                        "disponivel": "SIM",
                        "paciente": "",
                        "telefone": "",  # Formato: 53991189715 (DDD + número)
                        "status_confirmacao": ""  # PENDENTE, CONFIRMADO ou CANCELADO
                    })
    
    # Criar DataFrame
    df = pd.DataFrame(data)
    
    # Salvar com openpyxl para formatação
    df.to_excel(nome_arquivo, index=False, engine='openpyxl')
    
    # Formatar Excel
    wb = load_workbook(nome_arquivo)
    ws = wb.active
    
    # Estilos
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Formatar header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Ajustar largura das colunas
    larguras = {
        'A': 20,  # clinica
        'B': 15,  # exame
        'C': 12,  # data
        'D': 10,  # horario
        'E': 12,  # disponivel
        'F': 25,  # paciente
        'G': 15,  # telefone
        'H': 18   # status_confirmacao
    }
    
    for col, width in larguras.items():
        ws.column_dimensions[col].width = width
    
    # Formatar coluna telefone como TEXTO (evita notação científica)
    tel_col = 7  # Coluna G
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=tel_col)
        cell.number_format = '@'  # Formato texto
    
    # Salvar
    wb.save(nome_arquivo)
    
    print(f"\n✅ Planilha criada: {nome_arquivo}")
    print(f"📊 Total de horários: {len(df)}")
    print(f"🏥 Clínicas: {len(clinicas)}")
    print(f"👨‍⚕️ Especialidades: {len(exames)}")
    print(f"⏰ Horários por dia: {len(horarios)}")
    print(f"📅 Período: {df['data'].iloc[0]} até {df['data'].iloc[-1]}")
    
    print("\n📋 ESTRUTURA DA PLANILHA:")
    print("-" * 40)
    print("• clinica: Nome da clínica/UBS")
    print("• exame: Especialidade médica")
    print("• data: Data da consulta")
    print("• horario: Horário da consulta")
    print("• disponivel: SIM ou NAO")
    print("• paciente: Nome do paciente")
    print("• telefone: DDD + número (ex: 53991189715)")
    print("• status_confirmacao: PENDENTE/CONFIRMADO/CANCELADO")
    
    print("\n🔄 FLUXO DO SISTEMA:")
    print("-" * 40)
    print("1. Operador cadastra paciente na dashboard")
    print("2. Sistema marca 'disponivel' = NAO")
    print("3. Sistema salva nome e telefone")
    print("4. Sistema marca 'status_confirmacao' = PENDENTE")
    print("5. WhatsApp enviado para paciente")
    print("6. Paciente responde '1' (confirmar) ou '2' (cancelar)")
    print("7. Sistema atualiza 'status_confirmacao'")
    print("8. Se cancelar: 'disponivel' volta para SIM")
    
    return nome_arquivo


if __name__ == "__main__":
    # Criar planilha principal
    criar_planilha_agendamentos("agenda_clinicas.xlsx")
    
    # Criar também uma cópia exemplo
    criar_planilha_agendamentos("planilha_exemplo.xlsx")
    
    print("\n" + "=" * 60)
    print("✅ PLANILHAS CRIADAS COM SUCESSO!")
    print("=" * 60)
